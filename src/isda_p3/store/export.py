"""Box-7 Excel export (chunk 4.2) — a PRESENTATION view over the dataset.

The ``.xlsx`` is a *presentation* artifact for the slide / a reviewer's eyeballs.
Excel stores numbers as IEEE-754 floats, so values written here are floats — that
is acceptable *for display only*. The audited source of truth remains
``data/dataset/values.parquet``, where every figure persists as ``decimal128``
(see :mod:`isda_p3.store.dataset`). Nothing reads back from these workbooks.

What the export must NOT lose is provenance: every figure carries its
``source_url`` and ``page`` columns so a number on a slide can always be traced to
its source cell (CLAUDE.md §A.2). Empty data still yields a valid workbook with the
header row only — never a crash, never a silently absent sheet.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .analytics import PeerRow, TrendPoint, peer_compare, trend

#: Column order for a peer-comparison sheet. ``source_url`` + ``page`` are the
#: mandatory provenance columns (CLAUDE.md §A.2).
PEER_COLUMNS: tuple[str, ...] = (
    "bank",
    "value",
    "unit",
    "status",
    "confidence",
    "source_url",
    "page",
)

#: Column order for a trend sheet (one bank, one field, across periods).
TREND_COLUMNS: tuple[str, ...] = (
    "period",
    "value",
    "unit",
    "status",
    "confidence",
    "source_url",
    "page",
    "delta",
)

# The seven characters Excel forbids in a sheet title: [ ] : * ? / \ . The raw
# string r"[]:*?/\\" holds exactly those (the doubled backslash is one literal \).
_ILLEGAL_TITLE = set(r"[]:*?/\\")
_MAX_TITLE = 31

# A heterogeneous spreadsheet row (one cell per column), pre-serialisation.
_Row = list[str | float | int | None]


def _safe_title(name: str) -> str:
    """Sanitise a field code into a legal Excel sheet title (illegal chars -> ``_``)."""
    cleaned = "".join("_" if ch in _ILLEGAL_TITLE else ch for ch in name)
    return cleaned[:_MAX_TITLE] or "sheet"


def _write_header(ws: Worksheet, columns: Sequence[str]) -> None:
    """Write a bold header row and freeze it so it stays visible while scrolling."""
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _peer_row(row: PeerRow) -> _Row:
    """Flatten a :class:`PeerRow` to a cell list in :data:`PEER_COLUMNS` order.

    ``value``/``confidence`` become floats — this is the presentation boundary where
    Decimal -> float is deliberate and documented (the parquet store keeps Decimal).
    """
    return [
        row.bank,
        float(row.value),
        row.unit.value,
        row.status.value,
        float(row.confidence),
        row.source_url,
        row.page,
    ]


def _trend_row(point: TrendPoint) -> _Row:
    """Flatten a :class:`TrendPoint` to a cell list in :data:`TREND_COLUMNS` order."""
    return [
        point.period,
        float(point.value),
        point.unit.value,
        point.status.value,
        float(point.confidence),
        point.source_url,
        point.page,
        None if point.delta is None else float(point.delta),
    ]


def export_peer_xlsx(field_codes: Sequence[str], period: str, out_path: Path) -> Path:
    """Write a peer-comparison workbook: one sheet per ``field_code`` for ``period``.

    Each sheet lists every bank's value for that field (via
    :func:`isda_p3.store.analytics.peer_compare`, value-descending) with provenance
    columns. A field with no matching rows still gets a sheet with just the header
    (valid workbook, no crash). Raises ``FileNotFoundError`` (from ``peer_compare``)
    if the dataset file does not exist — "no file" is distinct from "no rows".
    Returns ``out_path``. Raises ``ValueError`` on empty ``field_codes`` — an empty
    request would otherwise produce a sheet-less, unsaveable workbook (a confusing
    crash); failing loud says exactly what was wrong.
    """
    if not field_codes:
        raise ValueError("field_codes is empty: nothing to export")
    wb = Workbook()
    wb.remove(wb.active)  # drop the auto-created blank sheet; we add named ones
    for code in field_codes:
        ws = wb.create_sheet(title=_safe_title(code))
        _write_header(ws, PEER_COLUMNS)
        for row in peer_compare(code, period):
            ws.append(_peer_row(row))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_trend_xlsx(bank_id: str, field_code: str, out_path: Path) -> Path:
    """Write a trend workbook: one sheet of ``bank_id``'s ``field_code`` over periods.

    Period-ascending with a period-over-period ``delta`` column (via
    :func:`isda_p3.store.analytics.trend`). No matching rows -> header only. Raises
    ``FileNotFoundError`` if the dataset is absent. Returns ``out_path``.
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=_safe_title(field_code))
    _write_header(ws, TREND_COLUMNS)
    for point in trend(bank_id, field_code):
        ws.append(_trend_row(point))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path

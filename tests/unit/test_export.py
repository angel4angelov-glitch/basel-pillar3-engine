"""Tests for isda_p3.store.export (chunk 4.2) — Excel presentation export.

The .xlsx is a *presentation* artifact (Excel stores floats); ``values.parquet``
remains the Decimal source of truth. What this gate protects is auditability, not
Decimal exactness: every figure exported MUST still carry its provenance, so the
``source_url`` and ``page`` columns are asserted present on every sheet (CLAUDE.md
§A.2 — no number without provenance). Empty data must still yield a valid workbook
(headers only), never a crash.
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest

from isda_p3.config import Paths
from isda_p3.models import (
    BBox,
    Bank,
    CheckType,
    EclBasis,
    Engine,
    FieldValue,
    FloorBasis,
    Jurisdiction,
    MappingDecision,
    MappingMethod,
    Provenance,
    ReconciliationResult,
    ReportingPeriod,
    SourceKind,
    Template,
    Unit,
    ValidationStatus,
)
from isda_p3.store.dataset import append_rows
from isda_p3.store.export import (
    PEER_COLUMNS,
    TREND_COLUMNS,
    export_peer_xlsx,
    export_trend_xlsx,
)

# --- builders (mirror test_analytics) ------------------------------------------

_MAP = MappingDecision(
    method=MappingMethod.RULE,
    model=None,
    prompt_sha=None,
    prompt_version=None,
    matched_alias="cet1 ratio",
    confidence=Decimal("1"),
)
_RUN_ID = "run-001"
_AT = "2026-06-19T00:00:00Z"


def _bank(bank_id: str, currency: str = "GBP") -> Bank:
    return Bank(
        id=bank_id,
        name=bank_id.title(),
        jurisdiction=Jurisdiction.UK,
        ir_url=f"https://{bank_id}.example",
        p3dh_lei=None,
        number_locale="en_GB",
        reporting_currency=currency,
    )


def _result(
    bank_id: str,
    period: ReportingPeriod,
    code: str,
    value: str,
    *,
    unit: Unit = Unit.PERCENT,
    confidence: str = "0.99",
    page: int | None = 4,
    status: ValidationStatus = ValidationStatus.AUTO_PASSED,
) -> ReconciliationResult:
    v = Decimal(value)
    bbox = None if page is None else BBox(page=page, x0=1.0, y0=2.0, x1=3.0, y1=4.0)
    prov = Provenance(
        bank_id=bank_id,
        period=period,
        source_url=f"https://{bank_id}.example/p3-{period.label}.pdf",
        source_kind=SourceKind.PDF,
        engine=Engine.DOCLING,
        bbox=bbox,
    )
    fv = FieldValue(
        template=Template.KM1,
        field_code=code,
        value=v,
        unit=unit,
        ecl_basis=EclBasis.NA,
        floor_basis=FloorBasis.FINAL,
        provenance=prov,
        mapping=_MAP,
        raw_text=value,
        engine_values={Engine.DOCLING: v},
    )
    return ReconciliationResult(
        field_value=fv,
        checks=(),
        confidence=Decimal(confidence),
        validation_basis=(CheckType.RATIO_IDENTITY,),
        status=status,
    )


@pytest.fixture
def tmp_dataset(tmp_path, monkeypatch):
    """Point the dataset store at an isolated tmp dir (mirrors test_analytics)."""
    dataset_dir = tmp_path / "dataset"
    monkeypatch.setattr(Paths, "DATASET_DIR", dataset_dir)
    monkeypatch.setattr(Paths, "DATASET", dataset_dir / "values.parquet")
    monkeypatch.setattr(
        Paths, "ensure", classmethod(lambda cls: dataset_dir.mkdir(parents=True, exist_ok=True))
    )
    return tmp_path


def _seed_peers(q: ReportingPeriod) -> None:
    append_rows([_result("barclays", q, "KM1.5", "13.6")], bank=_bank("barclays"),
                run_id=_RUN_ID, extracted_at=_AT)
    append_rows([_result("hsbc", q, "KM1.5", "14.8")], bank=_bank("hsbc"),
                run_id=_RUN_ID, extracted_at=_AT)


# --- export_peer_xlsx ----------------------------------------------------------


def test_export_peer_xlsx_sheet_header_and_value(tmp_dataset, tmp_path):
    _seed_peers(ReportingPeriod(2025, 4))
    out = export_peer_xlsx(["KM1.5"], "2025Q4", tmp_path / "peer.xlsx")
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert "KM1.5" in wb.sheetnames
    ws = wb["KM1.5"]

    header = [c.value for c in ws[1]]
    assert header == list(PEER_COLUMNS)
    # provenance columns are mandatory (CLAUDE.md §A.2)
    assert "source_url" in header
    assert "page" in header

    # rows are value-descending: hsbc 14.8 first, barclays 13.6 second
    bank_col = header.index("bank")
    value_col = header.index("value")
    body = list(ws.iter_rows(min_row=2, values_only=True))
    assert body[0][bank_col] == "hsbc"
    assert body[0][value_col] == 14.8  # float in the presentation export
    assert body[1][bank_col] == "barclays"
    assert body[1][value_col] == 13.6


def test_export_peer_xlsx_header_bold_and_frozen(tmp_dataset, tmp_path):
    _seed_peers(ReportingPeriod(2025, 4))
    out = export_peer_xlsx(["KM1.5"], "2025Q4", tmp_path / "peer.xlsx")
    ws = openpyxl.load_workbook(out)["KM1.5"]
    assert ws.freeze_panes == "A2"  # header row frozen
    assert all(c.font.bold for c in ws[1])  # header bold


def test_export_peer_xlsx_carries_provenance_values(tmp_dataset, tmp_path):
    _seed_peers(ReportingPeriod(2025, 4))
    out = export_peer_xlsx(["KM1.5"], "2025Q4", tmp_path / "peer.xlsx")
    ws = openpyxl.load_workbook(out)["KM1.5"]
    header = [c.value for c in ws[1]]
    body = list(ws.iter_rows(min_row=2, values_only=True))
    url_col, page_col = header.index("source_url"), header.index("page")
    for row in body:
        assert row[url_col].endswith("p3-2025Q4.pdf")
        assert row[page_col] == 4


def test_export_peer_xlsx_empty_data_writes_headers_only(tmp_dataset, tmp_path):
    _seed_peers(ReportingPeriod(2025, 4))
    # period with no matching rows: a valid workbook with just the header row.
    out = export_peer_xlsx(["KM1.5"], "2024Q4", tmp_path / "empty.xlsx")
    ws = openpyxl.load_workbook(out)["KM1.5"]
    assert [c.value for c in ws[1]] == list(PEER_COLUMNS)
    assert ws.max_row == 1  # header only, no data rows


def test_export_peer_xlsx_one_sheet_per_field(tmp_dataset, tmp_path):
    q = ReportingPeriod(2025, 4)
    append_rows([_result("barclays", q, "KM1.5", "13.6")], bank=_bank("barclays"),
                run_id=_RUN_ID, extracted_at=_AT)
    append_rows([_result("barclays", q, "KM1.6", "17.5")], bank=_bank("barclays"),
                run_id=_RUN_ID, extracted_at=_AT)
    out = export_peer_xlsx(["KM1.5", "KM1.6"], "2025Q4", tmp_path / "multi.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["KM1.5", "KM1.6"]


def test_export_peer_xlsx_empty_field_codes_raises(tmp_dataset, tmp_path):
    # An empty request would yield a sheet-less, unsaveable workbook: fail loud.
    _seed_peers(ReportingPeriod(2025, 4))
    with pytest.raises(ValueError, match="empty"):
        export_peer_xlsx([], "2025Q4", tmp_path / "x.xlsx")


def test_export_peer_xlsx_missing_dataset_raises(tmp_dataset, tmp_path):
    # "no file" is distinct from "no rows": a missing dataset must fail loud.
    with pytest.raises(FileNotFoundError):
        export_peer_xlsx(["KM1.5"], "2025Q4", tmp_path / "x.xlsx")


# --- export_trend_xlsx ---------------------------------------------------------


def test_export_trend_xlsx_periods_and_delta(tmp_dataset, tmp_path):
    series = {1: "13.1", 2: "13.3"}
    for qn in (1, 2):
        q = ReportingPeriod(2025, qn)
        append_rows([_result("barclays", q, "KM1.5", series[qn])],
                    bank=_bank("barclays"), run_id=_RUN_ID, extracted_at=_AT)
    out = export_trend_xlsx("barclays", "KM1.5", tmp_path / "trend.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "KM1.5" in wb.sheetnames
    ws = wb["KM1.5"]
    header = [c.value for c in ws[1]]
    assert header == list(TREND_COLUMNS)
    body = list(ws.iter_rows(min_row=2, values_only=True))
    period_col, delta_col = header.index("period"), header.index("delta")
    assert [r[period_col] for r in body] == ["2025Q1", "2025Q2"]
    assert body[0][delta_col] is None  # first point has no prior
    assert body[1][delta_col] == pytest.approx(0.2)


def test_export_trend_xlsx_empty_writes_headers_only(tmp_dataset, tmp_path):
    append_rows([_result("barclays", ReportingPeriod(2025, 1), "KM1.5", "13.1")],
                bank=_bank("barclays"), run_id=_RUN_ID, extracted_at=_AT)
    out = export_trend_xlsx("nobody", "KM1.5", tmp_path / "te.xlsx")
    ws = openpyxl.load_workbook(out)["KM1.5"]
    assert [c.value for c in ws[1]] == list(TREND_COLUMNS)
    assert ws.max_row == 1
